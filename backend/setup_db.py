"""
Create SQLite schema and import Excel data for DC Manager
"""
import sqlite3
import openpyxl
from datetime import datetime
import os
import sys

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '中联重科数据中心台账.xlsx')
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dc_manager.db')

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

# Create tables matching GORM models
cur.executescript('''
CREATE TABLE IF NOT EXISTS devices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source TEXT,
    asset_number TEXT,
    status TEXT,
    datacenter TEXT,
    cabinet TEXT,
    u_position TEXT,
    brand TEXT,
    model TEXT,
    device_type TEXT,
    serial_number TEXT,
    os TEXT,
    ip_address TEXT,
    system_account TEXT,
    mgmt_ip TEXT,
    mgmt_account TEXT,
    manufacture_date DATETIME,
    warranty_start DATETIME,
    warranty_end DATETIME,
    purpose TEXT,
    owner TEXT,
    remark TEXT,
    created_at DATETIME,
    updated_at DATETIME,
    deleted_at DATETIME
);

CREATE TABLE IF NOT EXISTS inspections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    device_id INTEGER,
    datacenter TEXT,
    cabinet TEXT,
    u_position TEXT,
    found_at DATETIME,
    inspector TEXT,
    assignee_id INTEGER,
    assignee_name TEXT,
    issue TEXT,
    severity TEXT,
    status TEXT,
    resolved_at DATETIME,
    escalation_level INTEGER DEFAULT 0,
    last_responded_at DATETIME,
    last_escalated_at DATETIME,
    remark TEXT,
    created_at DATETIME,
    updated_at DATETIME,
    deleted_at DATETIME,
    FOREIGN KEY (device_id) REFERENCES devices(id),
    FOREIGN KEY (assignee_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS inspection_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    inspection_id INTEGER NOT NULL,
    event_type TEXT NOT NULL,
    from_status TEXT,
    to_status TEXT,
    operator_id INTEGER,
    assignee_id INTEGER,
    assignee_name TEXT,
    escalation_level INTEGER DEFAULT 0,
    remark TEXT,
    webhook_status TEXT DEFAULT 'skipped',
    webhook_error TEXT,
    created_at DATETIME,
    FOREIGN KEY (inspection_id) REFERENCES inspections(id)
);
''')
conn.commit()
print("Schema created.")

# Import Excel
def to_str(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    s = str(v).strip()
    return '' if s.startswith('=ROW()') else s

def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(v.strip(), fmt).strftime('%Y-%m-%dT%H:%M:%SZ')
            except:
                pass
    return None

wb = openpyxl.load_workbook(XLSX_PATH)
total = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue

    header = [str(h).replace('\n', '').strip() if h else '' for h in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    def get(row, *names):
        for n in names:
            if n in col and col[n] < len(row):
                return to_str(row[col[n]])
        return ''

    def get_date(row, *names):
        for n in names:
            if n in col and col[n] < len(row):
                return to_date(row[col[n]])
        return None

    count = 0
    for row in rows[1:]:
        if not any(v for v in row if v is not None):
            continue
        brand = get(row, '设备\n品牌', '设备品牌')
        model_ = get(row, '设备型号')
        sn = get(row, '序列号')
        ip = get(row, 'IP地址')
        if not brand and not model_ and not sn and not ip:
            continue

        now = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
        cur.execute('''
            INSERT INTO devices (
                source, asset_number, status, datacenter, cabinet, u_position,
                brand, model, device_type, serial_number, os, ip_address,
                system_account, mgmt_ip, mgmt_account,
                manufacture_date, warranty_start, warranty_end,
                purpose, owner, remark, created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            sheet_name,
            get(row, '资产编号'),
            get(row, '状态'),
            get(row, '机房'),
            get(row, '机柜号', '新机柜号'),
            get(row, 'U位置', '设备位置\n（U数）', '设备位置（U数）'),
            brand, model_,
            get(row, '设备类型'),
            sn, get(row, '操作系统'), ip,
            get(row, '系统账号密码'),
            get(row, '远程管理IP'),
            get(row, '管理口账号'),
            get_date(row, '设备出厂时间'),
            get_date(row, '维保起始时间'),
            get_date(row, '维保结束时间'),
            get(row, '设备用途'),
            get(row, '责任人'),
            get(row, '备注说明', '描述'),
            now, now,
        ))
        count += 1

    conn.commit()
    print(f'  [{sheet_name}]: {count} devices')
    total += count

conn.close()
print(f'Done. Total: {total} devices imported to {DB_PATH}')
