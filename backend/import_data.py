"""
Import 中联重科数据中心台账.xlsx into SQLite dc_manager.db
"""
import sqlite3
import openpyxl
from datetime import datetime
import os
import sys

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '中联重科数据中心台账.xlsx')
DB_PATH = os.path.join(os.path.dirname(__file__), 'dc_manager.db')

if len(sys.argv) > 1:
    DB_PATH = sys.argv[1]
if len(sys.argv) > 2:
    XLSX_PATH = sys.argv[2]

def to_str(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return str(v).strip()

def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%dT%H:%M:%SZ')
    if isinstance(v, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(v.strip(), fmt).strftime('%Y-%m-%dT%H:%M:%SZ')
            except:
                pass
    return None

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

wb = openpyxl.load_workbook(XLSX_PATH)
total = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue

    header = [str(h).replace('\n', '').strip() if h else '' for h in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    def get(row, *names):
        for n in names:
            if n in col and col[n] < len(row):
                v = row[col[n]]
                s = to_str(v)
                if s and not s.startswith('=ROW()'):
                    return s
        return ''

    def get_date(row, *names):
        for n in names:
            if n in col and col[n] < len(row):
                v = row[col[n]]
                d = to_date(v)
                if d:
                    return d
        return None

    count = 0
    for i, row in enumerate(rows[1:], 2):
        if not any(v for v in row if v is not None):
            continue
        brand = get(row, '设备\n品牌', '设备品牌')
        model = get(row, '设备型号')
        sn = get(row, '序列号')
        ip = get(row, 'IP地址')
        if not brand and not model and not sn and not ip:
            continue

        now = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
        cur.execute('''
            INSERT INTO devices (
                source, asset_number, status, datacenter, cabinet, u_position,
                brand, model, device_type, serial_number, os, ip_address,
                system_account, mgmt_ip, mgmt_account,
                manufacture_date, warranty_start, warranty_end,
                purpose, owner, remark, created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            sheet_name,
            get(row, '资产编号'),
            get(row, '状态'),
            get(row, '机房'),
            get(row, '机柜号', '新机柜号'),
            get(row, 'U位置', '设备位置\n（U数）', '设备位置（U数）'),
            brand,
            model,
            get(row, '设备类型'),
            sn,
            get(row, '操作系统'),
            ip,
            get(row, '系统账号密码'),
            get(row, '远程管理IP'),
            get(row, '管理口账号'),
            get_date(row, '设备出厂时间'),
            get_date(row, '维保起始时间'),
            get_date(row, '维保结束时间'),
            get(row, '设备用途'),
            get(row, '责任人'),
            get(row, '备注说明', '描述'),
            now, now,
        ))
        count += 1

    conn.commit()
    print(f'Sheet [{sheet_name}]: imported {count} devices')
    total += count

conn.close()
print(f'Total imported: {total} devices')
